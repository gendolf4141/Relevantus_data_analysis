import time

import pandas as pd
from openpyxl.styles import PatternFill
import os
from openpyxl import load_workbook
from openpyxl.styles import Font


def add_mean_value(dataframe: pd.DataFrame) -> pd.DataFrame:
    value_for_group = dataframe.iloc[:,0].unique()
    result_df = pd.DataFrame()
    for value in value_for_group:
        df = dataframe[dataframe.iloc[:, 0] == value]
        mean_value = df.groupby(df.iloc[:, 0]).mean(numeric_only=True).reset_index()
        if df.shape[0] != 1:
            result_df = pd.concat([result_df, mean_value])
        result_df = pd.concat([result_df, df])
    return result_df


def re_spam(analiz_spam: list[list]) -> pd.DataFrame:
    columns = ['Слово (самая популярная словоформа)',
               'Повторы у Вас',
               'Минимум повторов (норм.)',
               'Максимум повторов (норм.)',
               'Переспам, %',
               'Переспам * IDF, %',
               'IDF',
               'Количество повторений',
               'URL']
    all_analiz_spam_df = pd.DataFrame()

    for url, file in analiz_spam:
        analiz_spam_df = pd.read_excel(file, sheet_name='Переспам')
        analiz_spam_df['URL'] = url
        all_analiz_spam_df = pd.concat([all_analiz_spam_df, analiz_spam_df])

    count_replay = all_analiz_spam_df.groupby('Слово (самая популярная словоформа)', as_index=False).agg(
        {'URL': 'count'}).rename(columns={'URL': 'Количество повторений'})

    all_analiz_spam_df = all_analiz_spam_df.merge(count_replay).sort_values(by=['Количество повторений',
                                                                                'Слово (самая популярная словоформа)'],
                                                                                ascending=False)
    all_dataframe = add_mean_value(all_analiz_spam_df[columns])

    return all_dataframe


def replay_word(replay_word: list[list]) -> pd.DataFrame:
    columns = ['Слово (самая популярная словоформа)',
               'Повторы у Вас',
               'Минимум повторов (норм.)',
               'Максимум повторов (норм.)',
               'Количество повторений',
               'URL']

    all_dataframe = pd.DataFrame()
    for url, file in replay_word:
        dataframe = pd.read_excel(file, sheet_name='Повторы слов')
        dataframe['URL'] = url
        all_dataframe = pd.concat([all_dataframe, dataframe])

    count_replay = all_dataframe.groupby('Слово (самая популярная словоформа)', as_index=False).agg(
        {'URL': 'count'}).rename(columns={'URL': 'Количество повторений'})

    all_dataframe = all_dataframe.merge(count_replay).sort_values(by=['Количество повторений',
                                                                                'Слово (самая популярная словоформа)'],
                                                                            ascending=False)
    all_dataframe = add_mean_value(all_dataframe[columns])
    return all_dataframe


def add_common_word(row_list: list[list]) -> pd.DataFrame:
    columns = ['Слово (самая популярная словоформа)',
               'Важные словоформы',
               'Все словоформы у конкурентов',
               'Количество повторений',
               'URL']

    all_dataframe = pd.DataFrame()
    for url, file in row_list:
        dataframe = pd.read_excel(file, sheet_name='Добавить важные слова')
        dataframe['URL'] = url
        all_dataframe = pd.concat([all_dataframe, dataframe])

    count_replay = all_dataframe.groupby('Слово (самая популярная словоформа)', as_index=False).agg(
        {'URL': 'count'}).rename(columns={'URL': 'Количество повторений'})

    all_dataframe = all_dataframe.merge(count_replay).sort_values(by=['Количество повторений',
                                                                      'Слово (самая популярная словоформа)'],
                                                                  ascending=False)
    all_dataframe = add_mean_value(all_dataframe[columns])

    return all_dataframe


def dop_word(row_list: list[list]) -> pd.DataFrame:
    columns = ['Дополнительные слова',
               'Количество повторений',
               'URL']

    all_dataframe = pd.DataFrame()
    for url, file in row_list:
        dataframe = pd.read_excel(file, sheet_name='Доп. слова')
        dataframe['URL'] = url
        all_dataframe = pd.concat([all_dataframe, dataframe])
    count_replay = all_dataframe.groupby('Дополнительные слова', as_index=False).agg({'URL': 'count'}).rename(columns={'URL': 'Количество повторений'})
    all_dataframe = all_dataframe.merge(count_replay).sort_values(by=['Количество повторений',
                                                                      'Дополнительные слова'],
                                                                  ascending=False)
    all_dataframe = add_mean_value(all_dataframe[columns])
    return all_dataframe


def title(row_list: list[list]) -> pd.DataFrame:
    columns = ['Можно добавить слова',
               'URL',
               'Количество повторений']

    all_dataframe = pd.DataFrame()
    for url, file in row_list:
        dataframe = pd.read_excel(file, sheet_name='title')
        dataframe['URL'] = url
        all_dataframe = pd.concat([all_dataframe, dataframe])

    count_replay = all_dataframe.groupby('Можно добавить слова', as_index=False).agg({'URL': 'count'}).rename(columns={'URL': 'Количество повторений'})
    all_dataframe = all_dataframe.merge(count_replay).sort_values(by=['Количество повторений',
                                                                      'Можно добавить слова'],
                                                                  ascending=False)
    all_dataframe = add_mean_value(all_dataframe[columns])
    return all_dataframe


def wight_row(path):
    wb = load_workbook(path)
    #
    # redFill = PatternFill(start_color='FFFF0000',
    #                       end_color='FFFF0000',
    #                       fill_type='solid')

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # размер шрифта документа
        font_size = 10
        # словарь с размерами столбцов
        cols_dict = {}
        list_null_value = []
        # проходимся по всем строкам документа
        for row in ws.rows:
            if row[-1].value is None:
                list_null_value.append(row[0].row)
            # теперь по ячейкам каждой строки
            for cell in row:
                # получаем букву текущего столбца
                letter = cell.column_letter
                name_letter = cell.coordinate
                if row[-1].value is None:
                    ws[name_letter].fill = PatternFill('solid', fgColor="DDDDDD")
                # если в ячейке записаны данные
                if cell.value:
                    # устанавливаем в ячейке размер шрифта
                    cell.font = Font(name='Calibri', size=font_size)
                    # вычисляем количество символов, записанных в ячейку
                    len_cell = len(str(cell.value))
                    # длинна колонки по умолчанию, если буква
                    # текущего столбца отсутствует в словаре `cols_dict`
                    len_cell_dict = 0
                    # смотрим в словарь c длинами столбцов
                    if letter in cols_dict:
                        # если в словаре есть буква текущего столбца
                        # то извлекаем соответствующую длину
                        len_cell_dict = cols_dict[letter]

                    # если текущая длина данных в ячейке
                    #  больше чем длинна из словаря
                    if len_cell > len_cell_dict:
                        # записываем новое значение ширины этого столбца
                        cols_dict[letter] = len_cell
                        ###!!! ПРОБЛЕМА АВТОМАТИЧЕСКОЙ ПОДГОНКИ !!!###
                        ###!!! расчет новой ширины колонки (здесь надо подгонять) !!!###
                        new_width_col = len_cell * font_size ** (font_size * 0.009)
                        # применение новой ширины столбца
                        ws.column_dimensions[cell.column_letter].width = new_width_col


        for count in range(len(list_null_value) - 1):
            min_row = int(list_null_value[count])
            max_row = int(list_null_value[count+1] - 2)
            ws.row_dimensions.group(min_row, max_row, hidden=True)

    wb.save(path)


def main():
    PATH = input("Введите путь: ")
    print('Формирую отчет...')
    # PATH = r'C:\Users\Gennady\Documents\Relevantus_data_analysis\files\Report_14_02_2023__21_55'
    FILE_RESULT = os.path.join(PATH, 'Result.xlsx')
    RELEVANTUS_DATA_ANALYSIS = os.path.join(PATH, 'Relevantus_data_analysis.xlsx')
    result_df = pd.read_excel(FILE_RESULT, sheet_name='Результаты')
    analiz_spam = result_df[['URL', 'Анализ переспама']].values.tolist()
    recommendations = result_df[['URL', 'Рекомендации по улучшению релевантности']].values.tolist()

    with pd.ExcelWriter(RELEVANTUS_DATA_ANALYSIS, engine='xlsxwriter') as writer:
        re_spam(analiz_spam).to_excel(writer, sheet_name='Переспам', index=False)
        replay_word(recommendations).to_excel(writer, sheet_name='Повторы слов', index=False)
        add_common_word(recommendations).to_excel(writer, sheet_name='Добавить важные слова', index=False)
        dop_word(recommendations).to_excel(writer, sheet_name='Доп. слова', index=False)
        title(recommendations).to_excel(writer, sheet_name='Title', index=False)

    wight_row(RELEVANTUS_DATA_ANALYSIS)
    print(f'Отчет сформирован, доступен по следующему пути: {RELEVANTUS_DATA_ANALYSIS}')
    time.sleep(5)


if __name__ == '__main__':
    main()